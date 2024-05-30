import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Load data from a CSV file
scheduled_hours_vs_actual_hours_df = pd.read_csv('Input files/scheduled_vs_actual_GT.csv')

# Define functions to manipulate dates
def convert_date_format(date_str):
    date_object = datetime.strptime(date_str, "%b %d, %Y")
    return date_object.strftime("%d-%b-%y")

def get_payroll_window(input_date):
    date = datetime.strptime(input_date, "%d-%b-%y")
    payroll_duration = 14
    days_since_start_of_period = (date - datetime(2024, 4, 21)).days % payroll_duration
    start_date = date - timedelta(days=days_since_start_of_period)
    end_date = start_date + timedelta(days=payroll_duration - 1)
    return start_date.strftime("%m.%d"), end_date.strftime("%m.%d")

def determine_payroll_period(date_str):
    start_date, end_date = get_payroll_window(date_str)
    return f"{start_date} - {end_date}"

def determine_week(date_str):
    date_object = datetime.strptime(date_str, "%d-%b-%y")
    week_num = (date_object.day - 1) // 7 + 1
    return f"{date_object.strftime('%b')} {week_num}w"

# Process the DataFrame
df = scheduled_hours_vs_actual_hours_df
df['Date'] = df['Date'].apply(convert_date_format)
df['Payroll Period'] = df['Date'].apply(determine_payroll_period)
df['Week'] = df['Date'].apply(determine_week)

# Save to Excel without index
filename = 'Generated files/Raw.xlsx'
df.to_excel(filename, index=False)

# Load the workbook and worksheet to apply formatting
wb = load_workbook(filename)
ws = wb.active

# Apply green fill to column headers
green_fill = PatternFill(start_color='50C878', end_color='50C878', fill_type='solid')
for cell in ws[1]:  # assuming the first row contains the headers
    cell.fill = green_fill

# Auto-adjust columns' width and rows' height
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    adjusted_width = (max_length + 2) * 1.2  # add a little extra space
    ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

for row in ws.iter_rows():
    max_height = max(cell.value.count('\n') + 1 if cell.value and '\n' in str(cell.value) else 1 for cell in row)
    adjusted_height = max_height * 15  # default height
    ws.row_dimensions[row[0].row].height = adjusted_height

# Apply auto-filter to all columns
ws.auto_filter.ref = ws.dimensions

# Save the changes to the workbook
wb.save(filename)

