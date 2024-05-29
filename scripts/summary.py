import openpyxl
from openpyxl.styles import Alignment, PatternFill
import pandas as pd

# Read the CSV data
df = pd.read_csv('excel_automation/Input files/scheduled_vs_actual_GT.csv')


# Calculate total scheduled hours and actual hours for each employee
employee_totals = df.groupby('Name').agg(
    total_scheduled_hours=('Scheduled Hours', 'sum'),
    total_actual_hours=('Actual Hours', 'sum')
).reset_index()

# Calculate the difference in hours
employee_totals['difference_hours'] = employee_totals['total_actual_hours'] - employee_totals['total_scheduled_hours']

# Use groupby with transform to get the count of occurrences of each name and add it as the 'Availability' column
employee_totals['Availability'] = df.groupby('Name').size().reset_index(name='Availability')['Availability']

# Create a new Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Report"

# Setting up the headers
headers = ["Location", "Staff Total", "Providers Total", "All Staff Ratio (Provider:Staff)"]
ws.append(headers)

# Adding Germantown row
ws.append(["Germantown", 0, 0, "#DIV/0!"])

# Adding empty row for spacing
ws.append([])

# Adding Staff/Provider table headers
staff_provider_headers = ["Staff/Provider Name", "Sum of Scheduled Hours", "Sum of Actual Hours", "Days of Availability", "Difference"]
ws.append(staff_provider_headers)

# Adding rows with data from employee_totals DataFrame
for index, row in employee_totals.iterrows():
    ws.append([row['Name'], row['total_scheduled_hours'], row['total_actual_hours'], row['Availability'], row['difference_hours']])

# Adding the final sum row and formatting
grand_total_row = ["Grand Total", employee_totals['total_scheduled_hours'].sum(), employee_totals['total_actual_hours'].sum(),
                   employee_totals['Availability'].sum(), employee_totals['difference_hours'].sum()]
ws.append(grand_total_row)

# Formatting the last row and difference value
emerald_green_fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid")
for cell in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for c in cell:
        c.fill = emerald_green_fill
        if c.column == 5:  # Difference column
            c.number_format = '0.00'

# Formatting the Difference column to two decimal places
for row in ws.iter_rows(min_row=4, max_row=ws.max_row-1, min_col=5, max_col=5):
    for cell in row:
        cell.number_format = '0.00'

# Aligning headers to center
for cell in ws["1:1"]:
    cell.alignment = Alignment(horizontal="center")

# Coloring the specific headers in emerald green
for cell in ws["4:4"]:
    cell.fill = emerald_green_fill

# Coloring the first set of headers in green
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
for cell in ws["1:1"]:
    cell.fill = green_fill

# Adjusting column widths to fit the values
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Add auto filter only to the Staff/Provider table headers section
ws.auto_filter.ref = f"A4:E{len(employee_totals) + 4}"

# Save the workbook
wb.save("excel_automation/Generated files/SummaryReport.xlsx")
print("hello")