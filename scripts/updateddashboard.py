
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

# Load the raw data
raw_data_path = 'Generated files/Raw.xlsx'  # replace with your actual file path
raw_df = pd.read_excel(raw_data_path)

timeoff_data_path = 'Input files/timeoffrequest_report_GT.csv'  # replace with your actual file path
timeoff_df = pd.read_csv(timeoff_data_path)

unpaid_replacement = 'UL'
pto_replacement = 'PTO'

# Efficiently update timeoff_df using boolean indexing
timeoff_df.loc[timeoff_df['Assigned Time Off Policies'] == 'Unpaid', 'Assigned Time Off Policies'] = unpaid_replacement
timeoff_df.loc[timeoff_df['Assigned Time Off Policies'] == 'PTO - 160 Hour', 'Assigned Time Off Policies'] = pto_replacement

# Load the format file
format_file_path = 'format file/format.xlsx'  # replace with your actual format file path
wb = load_workbook(format_file_path)
ws = wb.active

# Convert dates to datetime format
timeoff_df['Start Date'] = pd.to_datetime(timeoff_df['Start Date'])
timeoff_df['End Date'] = pd.to_datetime(timeoff_df['End Date'])

# Apply strftime to format dates
timeoff_df['Start Date'] = timeoff_df['Start Date'].dt.strftime("%d-%b")
timeoff_df['End Date'] = timeoff_df['End Date'].dt.strftime("%d-%b")

# Extract unique names and dates from the raw data
names = raw_df['Name'].unique()
dates = pd.to_datetime(raw_df['Date'].unique())

# Helper function to find the anchor cell of merged ranges
def get_anchor_cell(ws, cell):
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell

# Ensure dates are sorted and continuous
min_date = min(dates)
max_date = max(dates)
all_dates = pd.date_range(start=min_date, end=max_date)

# Calculate days based on dates
days = [date.strftime("%A") for date in all_dates]

# Fill in the names starting from cell C7
for row_index, name in enumerate(names, start=7):
    cell = ws.cell(row=row_index, column=3)
    if cell.value is None:
        cell.value = name

# Fill in the names starting from cell CO-7
for row_index, name in enumerate(names, start=7):
    cell = ws.cell(row=row_index, column=93)
    if cell.value is None:
        cell.value = name

# Fill in the Provider/Staff starting from cell C7
for row_index, name in enumerate(names, start=7):
    cell = ws.cell(row=row_index, column=2)
    if cell.value is None:
        cell.value = ""

# Fill in the Provider/Staff starting from cell CP-7
for row_index, name in enumerate(names, start=7):
    cell = ws.cell(row=row_index, column=94)
    if cell.value is None:
        cell.value = ""

# Fill in the days starting from cell 4D
for col_index, day in enumerate(days, start=0):  # Start from column D
    cell = ws.cell(row=4, column=col_index * 2 + 4)  # Adjust column index
    if cell.value is None:
        cell.value = day

# Fill in the dates starting from cell 5D
for col_index, date in enumerate(all_dates, start=0):  # Start from column D
    cell = ws.cell(row=5, column=col_index * 2 + 4)  # Adjust column index
    if cell.value is None:
        cell.value = date.strftime("%d-%b-%y")

# weeks = raw_df['Week'].unique()
# for col_index, week in enumerate(weeks, start=0):  # Start from column D
#     cell = ws.cell(row=5, column=col_index * 2 + 72)  # Adjust column index
#     if cell.value is None:
#         cell.value = week

weeks = raw_df['Week'].unique()
for col_index in range(0, len(weeks) - 1, 2):  # Step by 2 to process pairs
    week1 = weeks[col_index]
    week2 = weeks[col_index + 1]
    cell = ws.cell(row=5, column=(col_index // 2) * 2 + 72)  # Adjust column index
    if cell.value is None:
        cell.value = f"{week1} - {week2}"

min_date = min_date.strftime("%d-%b")
max_date = max_date.strftime("%d-%b")
date_range = f"{min_date} - {max_date}"
cell = ws.cell(row=5, column=84)
if cell.value is None:
    cell.value = date_range

# Loop through each name and populate the worksheet
for name_index, name in enumerate(names, start=0):
    # Filter data for the current name
    name_df = raw_df[raw_df['Name'] == name]
    for date_index, date in enumerate(all_dates, start=0):
        # Filter data for the current date
        date_str = date.strftime("%d-%b-%y")
        date_df = name_df[name_df['Date'] == date_str]
        
        # Calculate the row index for SH and AH
        row_index = name_index + 7
        if row_index == 18:
            row_index = 18  # Skip row 18
        col_index = date_index * 2 + 4

        # Populate the "Assigned Time Off Policies" cell
        assigned_time_offs = timeoff_df[(timeoff_df['Name'] == name) &
                                        (timeoff_df['Start Date'] <= date.strftime("%d-%b")) &
                                        (timeoff_df['End Date'] >= date.strftime("%d-%b"))]
        if not assigned_time_offs.empty:
            assigned_time_off_str = ', '.join(assigned_time_offs['Assigned Time Off Policies'])
            assigned_time_cell = ws.cell(row=row_index, column=col_index)  # Adjust column index for time-off policies
            if assigned_time_cell.value is None:
                # print("Hi")
                assigned_time_cell.value = assigned_time_off_str
        
        # Scheduled Hours (SH) cell
        sh_cell = ws.cell(row=row_index, column=col_index)
        if sh_cell.value is None:
            sh_cell.value = date_df.iloc[0]['Scheduled Hours'] if not date_df.empty else ''
        
        # Actual Hours (AH) cell
        ah_cell = ws.cell(row=row_index, column=col_index + 1)
        if ah_cell.value is None:
            ah_cell.value = date_df.iloc[0]['Actual Hours'] if not date_df.empty else ''
       

        

# Save the workbook to a new file
formatted_data_path = 'Generated files/updated_dashboard.xlsx'  # replace with your desired file path
wb.save(formatted_data_path)

print(f"Data successfully merged and saved to {formatted_data_path}")