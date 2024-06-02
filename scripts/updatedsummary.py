import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import pandas as pd

# Read the CSV data
csv_file_path = 'Input files/scheduled_vs_actual_GT.csv'
df = pd.read_csv(csv_file_path)

# Calculate total scheduled hours and actual hours for each employee
employee_totals = df.groupby('Name').agg(
    total_scheduled_hours=('Scheduled Hours', 'sum'),
    total_actual_hours=('Actual Hours', 'sum')
).reset_index()

# Calculate the difference in hours
employee_totals['difference_hours'] = employee_totals['total_actual_hours'] - employee_totals['total_scheduled_hours']

# Use groupby with transform to get the count of occurrences of each name and add it as the 'Availability' column
employee_totals['Availability'] = df.groupby('Name').size().reset_index(name='Availability')['Availability']

# Load the existing Excel file
excel_file_path = 'format file/summary format.xlsx'  # Corrected file extension
wb = openpyxl.load_workbook(excel_file_path)
ws = wb.active

# Define styles
header_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
sub_header_fill = PatternFill(start_color="50C878", end_color="50C878", fill_type="solid")
grand_total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Fill in the Location details
ws['A2'].value = "Location"

# Modify the Location cell under the Location column
ws['A5'].value = "#REF!"

# Adding Staff/Provider table headers
staff_provider_headers = ["Staff/Provider", "Name", "Sum of Scheduled Hours", "Sum of Actual Hours", "Days of Availability", "Difference"]
for col_num, header in enumerate(staff_provider_headers, 1):  # Start column enumeration from 1
    cell = ws.cell(row=4, column=col_num, value=header)
    cell.fill = sub_header_fill
    cell.font = bold_font
    cell.alignment = center_alignment
    cell.border = thin_border

start_row = 5
for index, row in employee_totals.iterrows():
    if index < len(employee_totals) - 1:
        # Skip filling the "Staff/Provider" column with actual data for all rows except the last one
        ws.cell(row=start_row + index, column=1, value="#REF!")  
    else:
        # Fill the last row of column 1 with the desired value
        ws.cell(row=start_row + index, column=1, value="#REF!")  

    ws.cell(row=start_row + index, column=2, value=row['Name'])
    ws.cell(row=start_row + index, column=3, value=row['total_scheduled_hours'])
    ws.cell(row=start_row + index, column=4, value=row['total_actual_hours'])
    ws.cell(row=start_row + index, column=5, value=row['Availability'])
    ws.cell(row=start_row + index, column=6, value=row['difference_hours'])


    # Apply styles to each cell in the row
    for col_num in range(2, 7):  # Adjusted column range
        cell = ws.cell(row=start_row + index, column=col_num)
        cell.alignment = center_alignment
        cell.border = thin_border

# Adding the Grand Total row
grand_total_row = start_row + len(employee_totals)
ws.cell(row=grand_total_row, column=1, value="#REF!")  # Leave this cell empty
ws.cell(row=grand_total_row, column=2, value="Grand Total")
ws.cell(row=grand_total_row, column=3, value=employee_totals['total_scheduled_hours'].sum())
ws.cell(row=grand_total_row, column=4, value=employee_totals['total_actual_hours'].sum())
ws.cell(row=grand_total_row, column=5, value=employee_totals['Availability'].sum())
ws.cell(row=grand_total_row, column=6, value=employee_totals['difference_hours'].sum())

# Apply styles to Grand Total row
for col_num in range(2, 7):  # Adjusted column range, excluding the first column
    cell = ws.cell(row=grand_total_row, column=col_num)
    cell.fill = grand_total_fill
    cell.font = bold_font
    cell.alignment = center_alignment
    cell.border = thin_border
# Adjust column widths to fit the values
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Add auto filter
ws.auto_filter.ref = f"B4:F{len(employee_totals) + 4}"  # Adjusted column range

# Save the workbook
output_file_path = "Generated files/Updated_SummaryReport.xlsx"
wb.save(output_file_path)

print(f"Updated Excel file saved to: {output_file_path}")