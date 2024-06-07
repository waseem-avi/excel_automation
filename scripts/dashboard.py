import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

# Load the raw data
raw_data_path = 'excel_automation/Generated files/Raw.xlsx'  # replace with your actual file path
raw_df = pd.read_excel(raw_data_path)

timeoff_data_path = 'excel_automation/Input files/timeoffrequest_report_GT.csv'  # replace with your actual file path
timeoff_df = pd.read_csv(timeoff_data_path)

# Load the format file
format_file_path = 'excel_automation/Format file/format.xlsx'  # replace with your actual format file path
wb = load_workbook(format_file_path)
ws = wb.active

grouped_data = timeoff_df.groupby('Assigned Time Off Policies')

# Initialize a dictionary to store counts of PTO, BT, and UL
time_off_counts = {
    'PTO': 0,
    'BT': 0,
    'UL': 0
}

# Count the occurrences of PTO, BT, and UL in the grouped data
for name, group in grouped_data:
    if name in time_off_counts:
        time_off_counts[name] += len(group)

# Extract unique names and dates from the raw data
names = raw_df['Name'].unique()
dates = pd.to_datetime(raw_df['Date'].unique())

# Helper function to find the anchor cell of merged ranges
def get_anchor_cell(ws, cell):
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell

# Ensure dates are sorted and continuous
min_date = min(dates)
max_date = max(dates)
all_dates = pd.date_range(start=min_date, end=max_date)

# Calculate days based on dates
days = [date.strftime("%A") for date in all_dates]

# Fill in the names starting from cell C7
for row_index, name in enumerate(names, start=7):
    cell = ws.cell(row=row_index, column=3)
    if cell.value is None:
        cell.value = name

# Fill in the names starting from cell CO-7
for row_index, name in enumerate(names, start=7):
    cell = ws.cell(row=row_index, column=93)
    if cell.value is None:
        cell.value = name

# Fill in the Provider/Staff starting from cell C7
for row_index, name in enumerate(names, start=7):
    cell = ws.cell(row=row_index, column=2)
    if cell.value is None:
        cell.value = "Provider/Staff"

# Fill in the Provider/Staff starting from cell CP-7
for row_index, name in enumerate(names, start=7):
    cell = ws.cell(row=row_index, column=94)
    if cell.value is None:
        cell.value = "Provider/Staff"

# Fill in the days starting from cell 4D
for col_index, day in enumerate(days, start=0):  # Start from column D
    cell = ws.cell(row=4, column=col_index * 2 + 4)  # Adjust column index
    if cell.value is None:
        cell.value = day

# Fill in the dates starting from cell 5D
for col_index, date in enumerate(all_dates, start=0):  # Start from column D
    cell = ws.cell(row=5, column=col_index * 2 + 4)  # Adjust column index
    if cell.value is None:
        cell.value = date.strftime("%d-%b-%y")

# weeks = raw_df['Week'].unique()
# for col_index, week in enumerate(weeks, start=0):  # Start from column D
#     cell = ws.cell(row=5, column=col_index * 2 + 72)  # Adjust column index
#     if cell.value is None:
#         cell.value = week


weeks = raw_df['Week'].unique()
for col_index in range(0, len(weeks) - 1, 2):  # Step by 2 to process pairs
    week1 = weeks[col_index]
    week2 = weeks[col_index + 1]
    cell = ws.cell(row=5, column=(col_index // 2) * 2 + 72)  # Adjust column index
    if cell.value is None:
        cell.value = f"{week1} - {week2}"


min_date = min_date.strftime("%d-%b")
max_date = max_date.strftime("%d-%b")
date_range = f"{min_date} - {max_date}"
cell=ws.cell(row=5, column=84)
if cell.value is None:
        cell.value = date_range

# Loop through each name and populate the worksheet
for name_index, name in enumerate(names, start=0):
    # Filter data for the current name
    name_df = raw_df[raw_df['Name'] == name]
    for date_index, date in enumerate(all_dates, start=0):
        # Filter data for the current date
        date_str = date.strftime("%d-%b-%y")
        date_df = name_df[name_df['Date'] == date_str]
        
        # Calculate the row index for SH and AH
        row_index = name_index + 7
        if row_index == 18:
            row_index = 18  # Skip row 18
        col_index = date_index * 2 + 4
        
        # Scheduled Hours (SH) cell
        sh_cell = ws.cell(row=row_index, column=col_index)
        if sh_cell.value is None:
            sh_cell.value = date_df.iloc[0]['Scheduled Hours'] if not date_df.empty else 0
        
        # Actual Hours (AH) cell
        ah_cell = ws.cell(row=row_index, column=col_index + 1)
        if ah_cell.value is None:
            ah_cell.value = date_df.iloc[0]['Actual Hours'] if not date_df.empty else 0

for assigned_policy, group in grouped_data:
    for index, row in group.iterrows():
        name = row['Name']
        start_date = pd.to_datetime(row['Start Date']).strftime('%d-%b')
        end_date = pd.to_datetime(row['End Date']).strftime('%d-%b')
        
        # Find the corresponding row in raw_df based on the name and date
        matching_rows = raw_df[(raw_df['Name'] == name) & (raw_df['Date'].between(start_date, end_date))]
        
        for _, matching_row in matching_rows.iterrows():
            scheduled_hours = matching_row['Scheduled Hours']
            actual_hours = matching_row['Actual Hours']
            date_str = matching_row['Date'].strftime("%d-%b-%y")

            # Find the row number in the Excel sheet to update
            for excel_row in range(7, ws.max_row + 1):
                if ws.cell(row=excel_row, column=3).value == name:
                    for col_index, date in enumerate(all_dates, start=0):
                        if date_str == date.strftime("%d-%b-%y"):
                            sh_cell = ws.cell(row=excel_row, column=col_index * 2 + 4)
                            ah_cell = ws.cell(row=excel_row, column=col_index * 2 + 5)
                            if sh_cell.value is None:
                                sh_cell.value = scheduled_hours
                            if ah_cell.value is None:
                                ah_cell.value = actual_hours

# Update the counts of PTO, BT, and UL in the Excel sheet
for row in range(7, ws.max_row + 1):
    name = ws.cell(row=row, column=3).value
    if name in time_off_counts:
        ws.cell(row=row, column=90).value = time_off_counts['PTO']
        ws.cell(row=row, column=91).value = time_off_counts['BT']
        ws.cell(row=row, column=92).value = time_off_counts['UL']

# Save the workbook to a new file
formatted_data_path = 'excel_automation/Generated files/dashboard.xlsx'  # replace with your desired file path
wb.save(formatted_data_path)