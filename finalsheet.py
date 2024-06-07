import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill,Border,Side
from openpyxl.utils import get_column_letter
import pandas as pd
def create_replicated_format_excel(filename):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set column widths
    column_widths = [20, 20, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Merge cells for "Provider / Staff"
    ws.merge_cells('A1:A3')
    provider_staff = ws['A1']
    provider_staff.value = "Provider / Staff"
    provider_staff.font = Font(bold=True)
    provider_staff.alignment = Alignment(horizontal='center', vertical='center')
    provider_staff.fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light Cyan Green

    # Set values for "Day" and "Date"
    ws['B1'] = "Day"
    ws['B1'].font = Font(bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light Cyan Green

    ws['B2'] = "Date"
    ws['B2'].font = Font(bold=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light Cyan Green

    # Set value and formatting for "Schedule Hours SH & Actual Hours AH"
    schedule_hours = ws['B3']
    schedule_hours.value = "Schedule Hours SH & Actual Hours AH"
    schedule_hours.font = Font(bold=True)
    schedule_hours.alignment = Alignment(horizontal='center', vertical='center')
    schedule_hours.fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light Cyan Green

    # Add borders to all cells
    for row in ws.iter_rows():
        for cell in row:
            border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
            cell.border = border

    # Save the workbook
    wb.save(filename)

def append_merged_columns_with_dates(filename, start_date, periods):
    try:
        # Try to load an existing workbook
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        wb = openpyxl.Workbook()

    ws = wb.active

    # Generate the date range
    date_range = pd.date_range(start=start_date, periods=periods)

    # Determine the starting column for new data
    max_column = ws.max_column
    start_col = max_column + 1

    for i, date in enumerate(date_range):
        col1 = get_column_letter(start_col + i * 2)
        col2 = get_column_letter(start_col + i * 2 + 1)

        # Merge cells in the new columns
        ws.merge_cells(f'{col1}1:{col2}1')
        ws.merge_cells(f'{col1}2:{col2}2')

        # Add day name to the top row with light cyan green fill
        day_name = date.strftime('%a')  # Day name (e.g., Sun, Mon)
        cell1 = ws[f'{col1}1']
        cell1.value = day_name
        cell1.alignment = Alignment(horizontal='center', vertical='center')
        cell1.fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light Cyan Green

        # Add date to the second row with light cyan green fill
        date_str = date.strftime('%d %b')  # Date (e.g., 21 Apr)
        cell2 = ws[f'{col1}2']
        cell2.value = date_str
        cell2.alignment = Alignment(horizontal='center', vertical='center')
        cell2.fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light Cyan Green

        # Add values to the bottom row with light orange (SH) and light blue (AH) fills
        ws[f'{col1}3'] = 'SH'
        ws[f'{col2}3'] = 'AH'
        ws[f'{col1}3'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Light Orange
        ws[f'{col2}3'].fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue

        # Add borders to all cells
        for row in ws.iter_rows(min_row=1, max_row=3, min_col=start_col + i * 2, max_col=start_col + i * 2 + 1):
            for cell in row:
                border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
                cell.border = border

    # Save the workbook to a file
    wb.save(filename)
from openpyxl.utils import column_index_from_string
def generate_dynamic_structure(filename, n):
    try:
        # Try to load an existing workbook
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        wb = openpyxl.Workbook()

    ws = wb.active

    # Determine the starting column for new data
    max_column = ws.max_column
    start_col = max_column + 1

    start_cell = get_column_letter(start_col)
    end_cell = get_column_letter(start_col + 2 * n - 1)

    # Merge cells in the first row
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 2 * n - 1)

    # Add value to the merged cell in the first row
    merged_cell = ws.cell(row=1, column=start_col, value="Weekly Hours")
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light Cyan Green

    # Create cells in the second row by merging 2 columns each, n times
    for i in range(1, n + 1):
        start_cell = get_column_letter(start_col + 2 * (i - 1))
        end_cell = get_column_letter(start_col + 2 * i - 1)
        ws.merge_cells(start_row=2, start_column=column_index_from_string(start_cell), end_row=2,
                       end_column=column_index_from_string(end_cell))
        merged_cell = ws.cell(row=2, column=start_col + 2 * (i - 1), value=f"Cell {i}")
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light Cyan Green

    # Create cells in the third row alternating 'SH' and 'AH'
    for i in range(start_col, start_col + 2 * n):
        value = 'SH' if (i - start_col) % 2 == 0 else 'AH'
        cell = ws.cell(row=3, column=i, value=value)
        cell.fill = PatternFill(start_color="FFD700" if value == 'SH' else "ADD8E6", 
                                end_color="FFD700" if value == 'SH' else "ADD8E6", fill_type="solid")
        
        # Add border to each cell
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        cell.border = border

    # Save the workbook to a file
    wb.save(filename)

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

def generate_modified_structure(filename, n):
    try:
        # Try to load an existing workbook
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        wb = openpyxl.Workbook()

    ws = wb.active

    # Determine the starting column for new data
    max_column = ws.max_column
    start_col = max_column + 1

    # Merge cells in the first row
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3 * n - 1)

    # Add value to the merged cell in the first row
    merged_cell = ws.cell(row=1, column=start_col, value="Months")
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light Cyan Green

    # Create cells in the second row by merging 3 columns each, n times
    for i in range(n):
        start_cell = get_column_letter(start_col + 3 * i)
        end_cell = get_column_letter(start_col + 3 * i + 2)
        ws.merge_cells(start_row=2, start_column=column_index_from_string(start_cell), end_row=2, end_column=column_index_from_string(end_cell))
        merged_cell = ws.cell(row=2, column=column_index_from_string(start_cell), value=f"Cell {i + 1}")
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")  # Light Cyan Green

    # Create cells in the third row alternating 'SH', 'AH', and 'Difference'
    for i in range(n):
        cell_sh = ws.cell(row=3, column=start_col + 3 * i, value='SH')
        cell_ah = ws.cell(row=3, column=start_col + 3 * i + 1, value='AH')
        cell_diff = ws.cell(row=3, column=start_col + 3 * i + 2, value='Difference')

        # Fill colors for SH, AH, and Difference
        cell_sh.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Light Orange
        cell_ah.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue
        cell_diff.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Sky Blue
        
        # Add border to each cell
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        cell_sh.border = border
        cell_ah.border = border
        cell_diff.border = border

    # Save the workbook to a file
    wb.save(filename)


import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

def create_absence_count_structure(filename):
    try:
        # Try to load an existing workbook
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        wb = openpyxl.Workbook()

    ws = wb.active

    # Determine the starting column for new data
    max_column = ws.max_column
    start_col = max_column + 1

    # Merge cells in the first row for 'Absence Count'
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 2)
    merged_cell = ws.cell(row=1, column=start_col, value="Absence Count")
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set fill color for 'Absence Count' - light cyan green
    for i in range(3):
        start_cell = get_column_letter(start_col + i)
        cell = ws.cell(row=1, column=column_index_from_string(start_cell))
        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    # Create cells in the second and third row for 'PTO', 'BT', 'UL'
    # Set fill color for each type - light red, red, light brown
    colors = ["FFCCCC", "FF6666", "CC9966"]  # Light red, red, light brown
    for i, color in enumerate(colors):
        start_cell = get_column_letter(start_col + i)
        ws.merge_cells(start_row=2, start_column=column_index_from_string(start_cell), end_row=3, end_column=column_index_from_string(start_cell))
        value = ['PTO', 'BT', 'UL'][i]
        merged_cell = ws.cell(row=2, column=column_index_from_string(start_cell), value=value)
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Add border to each cell
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        merged_cell.border = border

    # Save the workbook to a file
    wb.save(filename)

def create_name_structure(filename):
    try:
        # Try to load an existing workbook
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        wb = openpyxl.Workbook()

    ws = wb.active

    # Determine the starting column for new data
    max_column = ws.max_column
    start_col = max_column + 1

    # Merge cells in the first three rows for 'Name'
    ws.merge_cells(start_row=1, start_column=start_col, end_row=3, end_column=start_col)
    merged_cell = ws.cell(row=1, column=start_col, value="Name")
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set fill color for 'Name' - light cyan green
    for i in range(3):
        cell = ws.cell(row=i+1, column=start_col)
        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        # Add border to each cell
        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
        cell.border = border

    # Save the workbook to a file
    wb.save(filename)


# Example of calling the function
create_replicated_format_excel("replicated_format.xlsx")
append_merged_columns_with_dates('replicated_format.xlsx', '2023-04-21', 14)
generate_dynamic_structure("replicated_format.xlsx", 3)
generate_modified_structure("replicated_format.xlsx", 2)
create_absence_count_structure("replicated_format.xlsx")
create_name_structure("replicated_format.xlsx")